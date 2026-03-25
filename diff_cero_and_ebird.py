from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
from datetime import datetime

CERO = {'x': 'confirmed', 'nb': 'non breeding', 'in': 'introduced', 'ex':'extinct', 'e': 'endemic', 
'x/nb': 'confirmed + non breeding', 'v': 'vagrant', 'un': 'unconfirmed', 'un(i)' : 'introduced not breeding'}

def extract_cero_data():
    wb = load_workbook("files/ecuador_lista-master-final_2025_v4.xlsx")
    ws = wb.active  # or wb["Sheet1"]
    cero_data = {}
    cero_unconfirmed = {}
    cero_introduced = {}
    for row in ws.iter_rows(min_row=2):
        key = row[18].value   # Column S (Clem Common Name)
        value = row[4].value # Column E
    
        if key is not None:
            cero_data[key] = CERO[value]
            if value == 'un':
               cero_unconfirmed[key] = CERO[value]
            if value == 'in':
               cero_introduced[key] = CERO[value]
    return cero_data, cero_unconfirmed, cero_introduced

def extract_observation_dates(url):
    ebird_data = {}
    ebird_checklists = {}
    website = requests.get(url)
    soup = BeautifulSoup(website.text, 'html.parser')

    for li in soup.find_all('li', class_='BirdList-list-list-item'):
        time_tag = li.find('time')
        is_sensitive = li.find('button', class_='Sensitive-badge') is not None
        if (time_tag and time_tag.has_attr('datetime')) or is_sensitive:
            ebird_name = li.get('id')
            species = li.find('span', class_='Species-common').get_text(strip=True)
            if 'sp.' in species or '/' in species or '(hybrid)' in species or 'undescribed' in species:
                continue
            is_exotic = li.find('svg', class_='Icon--exoticEscapee') is not None
            if is_exotic:
                continue
            is_provisional = li.find('svg', class_='Icon--exoticProvisional') is not None
            if is_provisional:
                continue

            is_naturalized = li.find('svg', class_='Icon--exoticNaturalized') is not None
            if is_naturalized:
                value = f'ebird OK (NATURALIZED)'
            elif time_tag:
                checklist = li.find('div', class_='Obs-date').find('a').get('href')
                ebird_checklists[species] = checklist
                value = f'ebird OK ({datetime.strptime(time_tag["datetime"], "%Y-%m-%d %H:%M")})'
            else:
                value = 'ebird OK'
            ebird_data[species] = value
    return ebird_data, ebird_checklists

ebird_data, ebird_checklists = extract_observation_dates('https://ebird.org/region/EC/bird-list')
cero_data, cero_unconfirmed, cero_introduced = extract_cero_data()

only_in_ebird = {k: ebird_data[k] for k in ebird_data.keys() - cero_data.keys()}
only_in_cero = {k: cero_data[k] for k in cero_data.keys() - ebird_data.keys()}

print(f'ONLY IN EBIRD ({len(only_in_ebird)})')
print(only_in_ebird)
print('')
print(f'ONLY IN CERO ({len(only_in_cero)})')
print(only_in_cero)

print('')
print('___ INTRODUCED IN THE CERO LIST__')
for species in cero_introduced:
   print(species, ' : ', ebird_data[species])
print('')
print('___UNCONFIRMED IN THE CERO LIST__')
for species in cero_unconfirmed:
   print(species, ' ; ', ebird_data[species] if species in ebird_data else 'NO DATA',' ; ', ebird_checklists[species] if species in ebird_checklists else 'NO DATA')
